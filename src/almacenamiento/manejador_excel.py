from pathlib import Path

from openpyxl import Workbook

from src.configuracion import Estado, ConfigSimulacion


def guardar_resultado(
    ruta: Path,
    enfermedad: str,
    config: ConfigSimulacion,
    estadisticas: dict[Estado, int],
    dias: int,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulación"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    datos = [
        ("Enfermedad", enfermedad),
        ("Población Total", config.num_individuos),
        ("Radio de Infección", config.radio_infeccion),
        ("Prob. de Infección", f"{config.prob_infeccion:.4f}"),
        ("Prob. de Inmunidad", f"{config.prob_inmunidad:.4f}"),
        ("Tiempo de Vida", config.tiempo_muerte),
        ("Días Transcurridos", dias),
        ("Individuos Sanos", estadisticas[Estado.SANO]),
        ("Individuos Infectados", estadisticas[Estado.INFECTADO]),
        ("Individuos Inmunes", estadisticas[Estado.INMUNE]),
        ("Individuos Muertos", estadisticas[Estado.MUERTO]),
    ]

    for i, (campo, valor) in enumerate(datos, start=1):
        ws.cell(row=i, column=1, value=campo)
        ws.cell(row=i, column=2, value=valor)

    wb.save(ruta)
